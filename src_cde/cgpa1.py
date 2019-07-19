import os
import tkinter as tk
from tkinter import *
from tkinter import ttk
import openpyxl as op
import tkinter.messagebox as msg


class cgpa1:

    
    def cg1(self):
        

        #upto=w.get()
        #print(upto)
        cgpa_index=[]
        sem1_index=[]
        sem2_index=[]
        sem3_index=[]
        sem4_index=[]
        sem5_index=[]
        sem6_index=[]
        sem7_index=[]
        sem8_index=[]
        path=os.environ["HOMEPATH"]
        os.chdir(path)
        os.chdir("Desktop\\gpa")
        wb0=op.load_workbook('sem1.xlsx')
        sh0=wb0['Sheet1']

        cnt=sh0.max_row

        for i in range(3,cnt+1):
            if (sh0.cell(row=i,column=3).value and sh0.cell(row=i+1,column=3).value)==None:
                break
        cnt=i



        c=op.load_workbook('cgpa_sheet.xlsx')
        sh01=c['Sheet1']

        for i in range(4,cnt+1):
            sh01.cell(row=i,column=1).value=str(sh0.cell(row=i,column=1).value)
            sh01.cell(row=i,column=2).value=sh0.cell(row=i,column=2).value

        c.save('cgpa_sheet.xlsx')



        import semester_1

        wb0=op.load_workbook('sem1.xlsx')
        sh0=wb0['Sheet1']

        wb1=op.load_workbook('sem1.xlsx')
        sh1=wb1['Sheet1']
        for i in range(4,cnt+1):
            sem1_index.append(sh1.cell(row=i,column=sh1.max_column).value)
        #print(sem1_index)
        no_of_grades=len(sem1_index)
        #print(no_of_grades)
        cgpa_index=sem1_index

        
        wb=op.load_workbook('cgpa_sheet.xlsx')
        sh=wb['Sheet1']
        no_of_grades=len(cgpa_index)

        
        for i in range(4,cnt+1):

            sh.cell(row=i,column=3).value=sem1_index[i-4]

            sh.cell(row=i,column=4).value=cgpa_index[i-4]

        sh.cell(row=3,column=3).value='SEMESTER_1'
        sh.cell(row=3,column=4).value='CGPA'        

        if sh.cell(row=4,column=4).value!=None:
        
            msg.showinfo("SUCCESS","CGPA VALUES UPDATED IN 'cgpa_sheet' SHEET!!")

        else:
            msg.showinfo("SUCCESS","CGPA ERROR!!")

        wb.save('cgpa_sheet.xlsx')




obj=cgpa1()
obj.cg1()
