import os
import openpyxl as op
import tkinter.messagebox as msg
import pandas as pd
from pandas import ExcelFile
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import datetime
class semester_3:

    
    def sem3(self):
       

        path=os.environ["HOMEPATH"]
        os.chdir(path)
        os.chdir("Desktop\\gpa")
        wb=op.load_workbook('sem3.xlsx')
        sh=wb['Sheet1']
        #sh=wb.get_sheet_by_name('Sheet1')
        gr=['A','B','C','D','E','S']
        ab=['UA','WH','WH1','W','SA']
        pos=sh.max_column
        posr=sh.max_row
        for i in range(3,posr+1):
            if (sh.cell(row=i,column=3).value and sh.cell(row=i+1,column=3).value)==None:
                break
        posr=i
        #print(pos,posr)
        S=10
        A=9
        B=8
        C=7
        D=6
        E=5

        

        arrear=[]
        grade_index=[]
        for i in range(4,posr+1):
            acq_g=[]
            arr=0
            for j in range(3,sh.max_column+1):
                z=sh.cell(row=i,column=j).value
                acq_g.append(z)
            grade_index.append(acq_g)
        
            for v in acq_g:
                if v not in gr:
                    arr=arr+1
            arrear.append(arr)


        #print(arrear)
           

        gpa_index3=[]

        for i in range(4,posr+1):
            cp_index=[]
            bcp=0
            bcp_index=[]
            for j in range(3,pos+1):
        
                subno=j
                if subno==3:
                    bcp=4
                    bcp_index.append(bcp)
                elif subno==4:
                    bcp=3
                    bcp_index.append(bcp)
                elif subno==5:
                    bcp=3
                    bcp_index.append(bcp)
                elif subno==6:
                    bcp=3
                    bcp_index.append(bcp)
                elif subno==7:
                    bcp=3
                    bcp_index.append(bcp)
                elif subno==8:
                    bcp=3
                    bcp_index.append(bcp)
                elif subno==9:
                    bcp=2
                    bcp_index.append(bcp)
                elif subno==10:
                    bcp=2
                    bcp_index.append(bcp)
                

                

      
       

                if (grade_index[i-4][j-3])=='S':
                    cp=S*bcp
                    cp_index.append(cp)
                elif (grade_index[i-4][j-3])=='A':
                    cp=A*bcp
                    cp_index.append(cp)
                elif (grade_index[i-4][j-3])=='B':
                    cp=B*bcp
                    cp_index.append(cp)
                elif (grade_index[i-4][j-3])=='C':
                    cp=C*bcp
                    cp_index.append(cp)
                elif (grade_index[i-4][j-3])=='D':
                    cp=D*bcp
                    cp_index.append(cp)
                elif (grade_index[i-4][j-3])=='E':
                    cp=E*bcp
                    cp_index.append(cp)
                else:
                    cp=0
                    cp_index.append(cp)
            #print(cp_index)
            num=sum(cp_index)
            #print(num)
            #print(bcp_index)
            deno=sum(bcp_index)
            #print(deno)
            gpa=float(num/deno)
            #print('%.2f'%gpa)
            gpa_index3.append(float('%.2f'%gpa))
            #print(gpa_index3)

        

        #----------column cals-----
        stdcnt=posr-3
        #print(stdcnt)

        subarr_index=[]
        absent_index=[]
        gr_index=[]

        for i in range(3,pos+1):
            subarr=0
            
            for r in range(4,posr+1):
                
                if sh.cell(row=r,column=i).value not in gr:
                    subarr=subarr+1

            subarr_index.append(subarr)


        
        for i in range(3,pos+1):
            gra=0
            
            for r in range(4,posr+1):
                
                if sh.cell(row=r,column=i).value  in gr:
                    gra=gra+1

            gr_index.append(gra)




        for i in range(3,pos+1):
            absent=0
            for r in range(4,posr+1):
                if sh.cell(row=r,column=i).value in ab:
                    absent=absent+1
                    
            
            absent_index.append(absent)
            
        #print(subarr_index)
        #print(absent_index)

        appearence_index=[]

        for i in range(3,pos+1):
            appearence_index.append(stdcnt-absent_index[i-3])

        #print(appearence_index)
            

        arrpercent_index=[]

        for v in range (3,pos+1):
            if v==0:
                arrpercent=100
            else:
                arrpercent=(gr_index[v-3]/appearence_index[v-3])*100
            emptyval=float(arrpercent)
            arrpercent_index.append('%.2f'%emptyval)

        #print(arrpercent_index)

        nopass_index=[]
        nofail_index=[]
        for i in range(3,pos+1):
            nopass=0
            for r in range(4,posr+1):
                if sh.cell(row=r,column=i).value in gr:
                    nopass=nopass+1

            nopass_index.append(nopass)
            nofail_index.append(appearence_index[i-3]-nopass)
            
            


        #print(nopass_index)
        #print(nofail_index)
                   

        for i in range(4,posr+1):
            sh.cell(row=i,column=pos+2).value=gpa_index3[i-4]


        for i in range(4,posr+1):
            #al=arrear[i-4]
            sh.cell(row=i,column=pos+1).value=arrear[i-4]

        ovr=0
        for v in arrear:
                if v ==0:
                    ovr=ovr+1

        whcnt=0
        for i in range (4,posr+1):
            if sh.cell(row=i,column=3).value in ab and sh.cell(row=i,column=3).value!='UA':
                whcnt=whcnt+1
        #print(whcnt)
        without=stdcnt-whcnt
        #print(without)

        ovrpercent=float((ovr/without)*100)
        
                    


        if sh.cell(row=4,column=pos+2).value!=None:
            msg.showinfo("SUCCESS","GPA VALUES UPDATED FOR SEMESTER_3 IN 'sem3' SHEET!!")



        for i in range(3,pos+1):
            sh.cell(row=posr+2,column=i).value=stdcnt
            sh.cell(row=posr+3,column=i).value=absent_index[i-3]
            sh.cell(row=posr+4,column=i).value=appearence_index[i-3]
            sh.cell(row=posr+5,column=i).value=nofail_index[i-3]
            sh.cell(row=posr+6,column=i).value=nopass_index[i-3]
            sh.cell(row=posr+7,column=i).value=float(arrpercent_index[i-3])
            
            

        
    
        sh.cell(row=posr+2,column=1).value='TOTAL NO OF STUDENTS'
        sh.cell(row=posr+3,column=1).value='NO OF ABSENT STUDENT'
        sh.cell(row=posr+4,column=1).value='NO OF STUDENTS APPEARED'
        sh.cell(row=posr+5,column=1).value='SCORED LESS THAN 50%'
        sh.cell(row=posr+6,column=1).value='SCORED MORE THAN 50%'
        sh.cell(row=posr+7,column=1).value='PASS PERCENTAGE'
        sh.cell(row=posr+8,column=1).value='OVERALL PASS PERCENTAGE'
        sh.cell(row=posr+8,column=3).value=float('%.2f'%ovrpercent)

        

        sh.cell(row=3,column=pos+1).value='NO_OF_ARREARS'
        sh.cell(row=3,column=pos+2).value='GPA'
        sh.cell(row=2,column=1).value=str(datetime.datetime.now())

        wb.save('sem3.xlsx')


        #-------------------------------graph--------------------------------------

        g1=pd.read_excel('sem3.xlsx',sheet_name="Sheet1",skiprows=2)

        cols=g1.columns
        cols=(cols[2:-2]).tolist()

        #print(cols)

        
        arrpercent_index=list(map(float,arrpercent_index))
        #print(type(arrpercent_index[1]))
        #print(cols)
        rlis=pd.DataFrame({'Subject':cols,'percentage':arrpercent_index})



        rlis1=pd.DataFrame({'Subject':['OVR'],'percentage':ovrpercent})
        rlis=rlis.append(rlis1,ignore_index=True)
                

        
        #print(rlis)
        g=rlis.plot(kind='bar',title="3rd SEM analysis",x='Subject',y='percentage')
                        
        g.set_xlabel("SUBJECT")
        g.set_ylabel("PERCENTAGE")

        

        for i in g.patches:
            
            plt.annotate(str(i.get_height()),xy=i.get_xy(),xytext=(i.get_x(),(i.get_height()+0.5)))

        
        #plt.show()
        matplotlib.figure.SubplotParams(bottom=0.22,top=0.90)
        plt.subplots_adjust(bottom=0.22,top=0.90)
        fig=g.get_figure()
        

            
        path=os.environ["HOMEPATH"]
        os.chdir(path)
        os.chdir("Desktop\\gpa\\graphs")

        fig.savefig("g_sem3.jpg")

        
            
        path=os.environ["HOMEPATH"]
        os.chdir(path)
        os.chdir("Desktop\\gpa")


#--------------------------------------------------------------------------

obj=semester_3()
obj.sem3()


        
