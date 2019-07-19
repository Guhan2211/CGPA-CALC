import os
import tkinter as tk
from tkinter import *
from tkinter import ttk
import openpyxl as op
import tkinter.messagebox as msg
from resizeimage import resizeimage
from PIL import ImageTk
from PIL import Image as imk
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
import matplotlib.pyplot as plt
#matplotlib.use('TKAgg')


path=os.environ["HOMEPATH"]
os.chdir(path)

os.chdir("Desktop\\gpa")



img = imk.open(r'icon_cgpa.jpg')
img = resizeimage.resize_contain(img, [125,125])

class main(Frame):

    def __init__(self,master):
        super().__init__(master)

        var=tk.IntVar()

        lftframe=Frame(scr,bg='#00CED1',highlightbackground='WHITE',highlightthickness=2,padx=2)
        lftframe.grid(row=0,column=1)
        rtframe=Frame(scr,bg='#00CED1',padx=2)
        rtframe.grid(row=0,column=2)
        end_frame=Frame(scr,bg='#00CED1',highlightbackground='WHITE',highlightthickness=2,padx=2)
        end_frame.grid(row=0,column=3)





        l=tk.Label(lftframe,text="GPA",font=('veronica' ,20 ,'bold'),bg='#00CED1')
        l.grid(row=0,column=1,pady=6)

        b1=ttk.Button(lftframe,text="SEMESTER_1",command=self.s1)
        b1.grid(row=1,column=1,pady=12,padx=10)

        b2=ttk.Button(lftframe,text="SEMESTER_2",command=self.s2)
        b2.grid(row=2,column=1,pady=12,padx=10)

        b3=ttk.Button(lftframe,text="SEMESTER_3",command=self.s3)
        b3.grid(row=3,column=1,pady=12,padx=10)

        b4=ttk.Button(lftframe,text="SEMESTER_4",command=self.s4)
        b4.grid(row=4,column=1,pady=12,padx=10)

        b5=ttk.Button(lftframe,text="SEMESTER_5",command=self.s5)
        b5.grid(row=5,column=1,pady=12,padx=10)

        b6=ttk.Button(lftframe,text="SEMESTER_6",command=self.s6)
        b6.grid(row=6,column=1,pady=12,padx=10)

        b7=ttk.Button(lftframe,text="SEMESTER_7",command=self.s7)
        b7.grid(row=7,column=1,pady=12,padx=10)

        b8=ttk.Button(lftframe,text="SEMESTER_8",command=self.s8)
        b8.grid(row=8,column=1,pady=12,padx=10)




        l2=tk.Label(rtframe,text="CGPA",font=('veronica' ,20 ,'bold'),bg='#00CED1')
        l2.grid(row=0,column=0,pady=10)

        w = Spinbox(rtframe,values=(1,2,3,4,5,6,7,8),bg='white')
        w.grid(row=1,column=0,pady=5)




#-------------------------------cgpa------------------------



        def cg():
            upto=w.get()

            if upto=='1':
                import cgpa1
        
            elif upto=='2':
                import cgpa2
         
            elif upto=='3':
                import cgpa3

            elif upto=='4':
                import cgpa4

            elif upto=='5':
                import cgpa5

            elif upto=='6':
                import cgpa6

            elif upto=='7':
                import cgpa7

            elif upto=='8':
                import cgpa8




        b9=ttk.Button(rtframe,text="CALCULATE",command=cg)
        b9.grid(row=2,column=0,pady=10,padx=10)



#------------------------------ovr graph--------------------------------------------
        l7=tk.Label(rtframe,text="OVERALL",font=('veronica' ,20 ,'bold'),bg='#00CED1')
        l7.grid(row=4,column=0,pady=10)


        path=os.environ["HOMEPATH"]
        os.chdir(path)

        os.chdir("Desktop\\gpa")


        

        ovr_val=StringVar()

        
        
        r1=tk.Radiobutton(rtframe,text="ODD SEMESTER",variable=ovr_val,value="O",bg='#00CED1')
        r2=tk.Radiobutton(rtframe,text="EVEN SEMESTER",variable=ovr_val,value="E",bg='#00CED1')
        r1.grid(row=5,column=0,padx=10)
        r2.grid(row=6,column=0,padx=30)


        def graph():
            sdf=ovr_val.get()
            print(sdf)

            if (sdf=="O"):

                xl1=pd.read_excel('sem1.xlsx',sheet_name="Sheet1",skiprows=2)
                xl3=pd.read_excel('sem3.xlsx',sheet_name="Sheet1",skiprows=2)
                xl5=pd.read_excel('sem5.xlsx',sheet_name="Sheet1",skiprows=2)
                xl7=pd.read_excel('sem7.xlsx',sheet_name="Sheet1",skiprows=2)

                xl1=xl1.drop(xl1.index[:-1])
                col1=xl1.columns
                r1=xl1.at[xl1.index[0],col1[2]]

                xl3=xl3.drop(xl3.index[:-1])
                col3=xl3.columns
                r3=xl3.at[xl3.index[0],col3[2]]

                xl5=xl5.drop(xl5.index[:-1])
                col5=xl5.columns
                r5=xl5.at[xl5.index[0],col5[2]]

                xl7=xl7.drop(xl7.index[:-1])
                col7=xl7.columns
                r7=xl7.at[xl7.index[0],col7[2]]

                
                print(r1,r3,r5,r7)
                res=(r1+r3+r5+r7)/4
                print(res)
                rlis=pd.DataFrame({'year':['ovr','1yr','2yr','3yr','4yr'],'percentage':[res,r1,r3,r5,r7]})
                
                print(rlis)
                g=rlis.plot(kind='bar',title="Overall CSE Performance",x='year',y='percentage')
                
                g.set_xlabel("YEAR")
                g.set_ylabel("PERCENTAGE")
                for i in g.patches:
                    #print(i)
                    #print(i.get_height())
                    #print(i.get_xy())
                    #print(i.get_x())
                    
                    plt.annotate(str(i.get_height()),xy=i.get_xy(),xytext=(i.get_x(),(i.get_height()+0.5)))
                plt.show()
                fig=g.get_figure()
                fig.savefig("result.jpg")

            if (sdf=="E"):
                                
                xl2=pd.read_excel('sem2.xlsx',sheet_name="Sheet1",skiprows=2)
                xl4=pd.read_excel('sem4.xlsx',sheet_name="Sheet1",skiprows=2)
                xl6=pd.read_excel('sem6.xlsx',sheet_name="Sheet1",skiprows=2)
                xl8=pd.read_excel('sem8.xlsx',sheet_name="Sheet1",skiprows=2)

                xl2=xl2.drop(xl2.index[:-1])
                col2=xl2.columns
                r2=xl2.at[xl2.index[0],col2[2]]

                xl4=xl4.drop(xl4.index[:-1])
                col4=xl4.columns
                r4=xl4.at[xl4.index[0],col4[2]]

                xl6=xl6.drop(xl6.index[:-1])
                col6=xl6.columns
                r6=xl6.at[xl6.index[0],col6[2]]

                xl8=xl8.drop(xl8.index[:-1])
                col8=xl8.columns
                r8=xl8.at[xl8.index[0],col8[2]]

               
                #print(r2,r4,r6,r8)
                res=(r2+r4+r6+r8)/4
                #print(res)


                #print(res)
                rlis=pd.DataFrame({'year':['ovr','1yr','2yr','3yr','4yr'],'percentage':[res,r2,r4,r6,r8]})
                
                #print(rlis)
                g=rlis.plot(kind='bar',title="Overall CSE Performance",x='year',y='percentage')
                
                g.set_xlabel("YEAR")
                g.set_ylabel("PERCENTAGE")

                for i in g.patches:
                    #print(i)
                    #print(i.get_height())
                    #print(i.get_xy())
                    #print(i.get_x())
                    
                    plt.annotate(str(i.get_height()),xy=i.get_xy(),xytext=(i.get_x(),(i.get_height()+0.5)))
                plt.show()
                fig=g.get_figure()
                path=os.environ["HOMEPATH"]
                os.chdir(path)

                os.chdir("Desktop\\gpa\\graphs")
                fig.savefig("result.jpg")


            msg.showinfo("OVR","PERFORMANCE IMAGE SAVED in GRAPHS folder!")


            
        

        b10=ttk.Button(rtframe,text="GET OVR",command=graph)
        b10.grid(row=7,column=0,pady=10,padx=10)
        

        

        
#------------------------------image_place------------------------------------------

        can=Canvas(rtframe,width=125,height=125,bg='#00CED1')
        can.grid(row=3,column=0,pady=5,padx=10)

        s1=ImageTk.PhotoImage(img)
        label=Label(image=s1)
        label.image=s1

        can.create_image((0,0),image=s1,anchor='nw')

#-------------------------------------------------------------------------------------

        l0=tk.Label(end_frame,text="ARREARS",font=('veronica' ,20 ,'bold'),bg='#00CED1')
        l0.grid(row=0,column=0,pady=10)

        

        rd_val=StringVar()

        
        
        r1=tk.Radiobutton(end_frame,text="SEMESTER-WISE",variable=rd_val,value="S",bg='#00CED1')
        r2=tk.Radiobutton(end_frame,text="CUMMULATIVE",variable=rd_val,value="C",bg='#00CED1')
        r1.grid(row=1,column=0,padx=10)
        r2.grid(row=2,column=0,padx=10)






        opt_val=tk.IntVar(scr)
        l=tk.Label(end_frame,text="For/Untill semester",font=('veronica' ,10 ,'bold'),bg='#00CED1')
        l.grid(row=3,column=0,pady=10,padx=10)

        x=ttk.OptionMenu(end_frame,opt_val,1,2,3,4,5,6,7,8)
        x.grid(row=4,column=0,pady=10,padx=10)

#----------------------------------------arrear_calcs_modules-----------------------------------------------------------
        def sg():
            crs=rd_val.get()
            #print(crs)
            numb=opt_val.get()
            #print(numb)
            #print(type(numb))

            path=os.environ["HOMEPATH"]
            os.chdir(path)

            os.chdir("Desktop\\gpa")

            wb=op.load_workbook('arrear.xlsx')
            l=wb.sheetnames

            for shee in l:
                sh=wb[shee]
	
                for i in range(1,sh.max_row+1):  
                    for j in range(1,sh.max_column+1):
                        sh.cell(row=i,column=j).value=None

            wb.save('arrear.xlsx')


                                    
            xl1=pd.read_excel('sem1.xlsx',sheet_name="Sheet1",skiprows=2)
            xl2=pd.read_excel('sem2.xlsx',sheet_name="Sheet1",skiprows=2)
            xl3=pd.read_excel('sem3.xlsx',sheet_name="Sheet1",skiprows=2)
            xl4=pd.read_excel('sem4.xlsx',sheet_name="Sheet1",skiprows=2)
            xl5=pd.read_excel('sem5.xlsx',sheet_name="Sheet1",skiprows=2)
            xl6=pd.read_excel('sem6.xlsx',sheet_name="Sheet1",skiprows=2)
            xl7=pd.read_excel('sem7.xlsx',sheet_name="Sheet1",skiprows=2)
            xl8=pd.read_excel('sem8.xlsx',sheet_name="Sheet1",skiprows=2)


            res=pd.read_excel('arrear.xlsx',sheet_name="Sheet1",skiprows=2)

            
            #------------------------dependency_calc-------------------------------------

            


            if (crs=='C' ):

                    if(numb==1):
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                           
                            
                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                           
                            
                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)
                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                    elif(numb==2):

                                            
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            
                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            
                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)

                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                    elif(numb==3):

                                            
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            xl3=xl3.drop(xl3.index[-8:])
                           
                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            
                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)

                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                            #3
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar3['NO_OF_ARREARS']))
                    elif(numb==4):


                                            
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            xl3=xl3.drop(xl3.index[-8:])
                            xl4=xl4.drop(xl4.index[-8:])
                            

                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar4=xl4[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            

                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)

                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                            #3
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar3['NO_OF_ARREARS']))
                            #4
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar4['NO_OF_ARREARS']))
                    elif(numb==5):

                        
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            xl3=xl3.drop(xl3.index[-8:])
                            xl4=xl4.drop(xl4.index[-8:])
                            xl5=xl5.drop(xl5.index[-8:])
                           

                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar4=xl4[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar5=xl5[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)




                            
                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                            #3
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar3['NO_OF_ARREARS']))
                            #4
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar4['NO_OF_ARREARS']))
                            #5
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar5['NO_OF_ARREARS']))
                    elif(numb==6):


                                            
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            xl3=xl3.drop(xl3.index[-8:])
                            xl4=xl4.drop(xl4.index[-8:])
                            xl5=xl5.drop(xl5.index[-8:])
                            xl6=xl6.drop(xl6.index[-8:])
                            

                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar4=xl4[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar5=xl5[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar6=xl6[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            

                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)

                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                            #3
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar3['NO_OF_ARREARS']))
                            #4
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar4['NO_OF_ARREARS']))
                            #5
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar5['NO_OF_ARREARS']))
                            #6
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar6['NO_OF_ARREARS']))
                    elif(numb==7):

                                            
                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            xl3=xl3.drop(xl3.index[-8:])
                            xl4=xl4.drop(xl4.index[-8:])
                            xl5=xl5.drop(xl5.index[-8:])
                            xl6=xl6.drop(xl6.index[-8:])
                            xl7=xl7.drop(xl7.index[-8:])
                            

                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar4=xl4[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar5=xl5[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar6=xl6[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar7=xl7[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            

                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)



                        
                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                            #3
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar3['NO_OF_ARREARS']))
                            #4
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar4['NO_OF_ARREARS']))
                            #5
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar5['NO_OF_ARREARS']))
                            #6
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar6['NO_OF_ARREARS']))
                            #7
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar7['NO_OF_ARREARS']))
                    elif(numb==8):

                            y=xl1.shape

                            xl1=xl1.drop(xl1.index[-8:])
                            xl2=xl2.drop(xl2.index[-8:])
                            xl3=xl3.drop(xl3.index[-8:])
                            xl4=xl4.drop(xl4.index[-8:])
                            xl5=xl5.drop(xl5.index[-8:])
                            xl6=xl6.drop(xl6.index[-8:])
                            xl7=xl7.drop(xl7.index[-8:])
                            xl8=xl8.drop(xl8.index[-8:])

                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar4=xl4[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar5=xl5[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar6=xl6[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar7=xl7[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            ar8=xl8[['REGISTER_NO','NAME','NO_OF_ARREARS']]

                            #print(ar1)

                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)

                        
                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            #2
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar2['NO_OF_ARREARS']))
                            #3
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar3['NO_OF_ARREARS']))
                            #4
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar4['NO_OF_ARREARS']))
                            #5
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar5['NO_OF_ARREARS']))
                            #6
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar6['NO_OF_ARREARS']))
                            #7
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar7['NO_OF_ARREARS']))
                            #8
                            res['NO_OF_ARREARS']=(res['NO_OF_ARREARS'].add(ar8['NO_OF_ARREARS']))
            #-------------------------------single sem arr sheet-----------------------
            if (crs=='S'):

                    if(numb==1):
                            y=xl1.shape
                            xl1=xl1.drop(xl1.index[-8:])                            
                            ar1=xl1[['REGISTER_NO','NAME','NO_OF_ARREARS']]
                            #print(ar1)
                            res['REGISTER_NO']=(ar1['REGISTER_NO']).astype(str)
                            res['NAME']=ar1['NAME']
                            #print(res)
                            
                            #1
                            res['NO_OF_ARREARS']=(ar1['NO_OF_ARREARS'])
                            
                    elif(numb==2):
                            y=xl2.shape
                            xl2=xl2.drop(xl2.index[-8:])                           
                            ar2=xl2[['REGISTER_NO','NAME','NO_OF_ARREARS']] 
                            #print(ar1)
                            res['REGISTER_NO']=(ar2['REGISTER_NO']).astype(str)
                            res['NAME']=ar2['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar2['NO_OF_ARREARS'])
                    elif(numb==3):

                            y=xl3.shape
                            xl3=xl3.drop(xl3.index[-8:])                                                       
                            ar3=xl3[['REGISTER_NO','NAME','NO_OF_ARREARS']]                                                       
                            #print(ar1)
                            res['REGISTER_NO']=(ar3['REGISTER_NO']).astype(str)
                            res['NAME']=ar3['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar3['NO_OF_ARREARS'])
                    elif(numb==4):

                            y=xl4.shape
                            xl4=xl4.drop(xl4.index[-8:])                           
                            ar4=xl4[['REGISTER_NO','NAME','NO_OF_ARREARS']] 
                            #print(ar1)
                            res['REGISTER_NO']=(ar4['REGISTER_NO']).astype(str)
                            res['NAME']=ar4['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar4['NO_OF_ARREARS'])
                    elif(numb==5):

                            y=xl5.shape
                            xl5=xl5.drop(xl5.index[-8:])                           
                            ar5=xl5[['REGISTER_NO','NAME','NO_OF_ARREARS']] 
                            #print(ar1)
                            res['REGISTER_NO']=(ar5['REGISTER_NO']).astype(str)
                            res['NAME']=ar5['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar5['NO_OF_ARREARS'])
                    elif(numb==6):
                            y=xl6.shape
                            xl6=xl6.drop(xl6.index[-8:])                           
                            ar6=xl6[['REGISTER_NO','NAME','NO_OF_ARREARS']] 
                            #print(ar1)
                            res['REGISTER_NO']=(ar6['REGISTER_NO']).astype(str)
                            res['NAME']=ar6['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar6['NO_OF_ARREARS'])
                    elif(numb==7):
                            y=xl7.shape
                            xl7=xl7.drop(xl7.index[-8:])                           
                            ar7=xl7[['REGISTER_NO','NAME','NO_OF_ARREARS']] 
                            #print(ar1)
                            res['REGISTER_NO']=(ar7['REGISTER_NO']).astype(str)
                            res['NAME']=ar7['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar7['NO_OF_ARREARS'])
                    elif(numb==8):

                            y=xl8.shape
                            xl8=xl8.drop(xl8.index[-8:])                           
                            ar8=xl8[['REGISTER_NO','NAME','NO_OF_ARREARS']] 
                            #print(ar1)
                            res['REGISTER_NO']=(ar8['REGISTER_NO']).astype(str)
                            res['NAME']=ar8['NAME']
                            #print(res)
                            
                            res['NO_OF_ARREARS']=(ar8['NO_OF_ARREARS'])




            #print(res)


            arr_df=pd.DataFrame(columns=['REGISTER_NO','NAME','NO_OF_ARREARS'])
            nilarr_df=pd.DataFrame(columns=['REGISTER_NO','NAME'])
            #print(np_df)

            #print(y[0])

            for i in range(0,y[0]-8):
                    l=[]
                    if (res.at[i,'NO_OF_ARREARS']!=0):
                            #print(res.at[i,'NO_OF_ARREARS'])
                            
                            l=[res.at[i,'REGISTER_NO'],res.at[i,'NAME'],(res.at[i,'NO_OF_ARREARS'])]
                            #print(l)

                            
                            arr_df.loc[len(arr_df)]=l
                            arr_df=arr_df.sort_values('NO_OF_ARREARS')

                    else:
                            l=[res.at[i,'REGISTER_NO'],res.at[i,'NAME']]
                            nilarr_df.loc[len(nilarr_df)]=l



            cnt=arr_df.shape
            one_arr_df=pd.DataFrame(columns=['REGISTER_NO','NAME'])
            two_arr_df=pd.DataFrame(columns=['REGISTER_NO','NAME'])
            mre_arr_df=pd.DataFrame(columns=['REGISTER_NO','NAME'])

            for i in range(0,cnt[0]+1):
                    
                    l=[]
                    if (res.at[i,'NO_OF_ARREARS']==1):
                            l=[res.at[i,'REGISTER_NO'],res.at[i,'NAME']]
                            one_arr_df.loc[len(one_arr_df)]=l
            for i in range(0,cnt[0]+1):
                    if (res.at[i,'NO_OF_ARREARS']==2):
                            l=[res.at[i,'REGISTER_NO'],res.at[i,'NAME']]
                            two_arr_df.loc[len(two_arr_df)]=l

            for i in range(0,cnt[0]+1):
                    if (res.at[i,'NO_OF_ARREARS']>2):
                            l=[res.at[i,'REGISTER_NO'],res.at[i,'NAME']]
                            mre_arr_df.loc[len(mre_arr_df)]=l


            '''print(arr_df)
            print('\n nill')
            print(nilarr_df)
            print('\n one')
            print(one_arr_df)
            print('\n two')
            print(two_arr_df)
            print('\n mre')
            print(mre_arr_df)'''

                                              


            writ=pd.ExcelWriter('arrear.xlsx',engine='xlsxwriter')
            arr_df.to_excel(writ,sheet_name="Sheet1")
            nilarr_df.to_excel(writ,sheet_name="NIL_ARREAR")
            one_arr_df.to_excel(writ,sheet_name="ONE_ARREAR")
            two_arr_df.to_excel(writ,sheet_name="TWO_ARREAR")
            mre_arr_df.to_excel(writ,sheet_name="MORE_ARREAR")

            msg.showinfo("SUCCESS","ARREAR LIST CREATED in 'arrear' SHEET!!")

            writ.save()



            

#--------------------------------------------------------------------------------------------------------------            
            

        bes=ttk.Button(end_frame,text="GENERATE",command=sg)
        bes.grid(row=5,column=0,pady=10,padx=10)

#-----------------------------TICK-FORMAT----------------------------------------------------------------

        l02=tk.Label(end_frame,text="TICK-FORMAT",font=('veronica' ,20 ,'bold'),bg='#00CED1')
        l02.grid(row=6,column=0,pady=10)

        rd_val2=StringVar()

        
        
        r11=tk.Radiobutton(end_frame,text="SEMESTER-WISE",variable=rd_val2,value="S",bg='#00CED1')
        r22=tk.Radiobutton(end_frame,text="CUMMULATIVE",variable=rd_val2,value="C",bg='#00CED1')
        r11.grid(row=7,column=0,padx=10)
        r22.grid(row=8,column=0,padx=10)




        opt_val5=tk.IntVar(scr)
        x5=ttk.OptionMenu(end_frame,opt_val5,1,2,3,4,5,6,7,8)
        x5.grid(row=9,column=0,pady=10,padx=10)



        def yg():
            tn=opt_val5.get()
            tn2=rd_val2.get()

            path=os.environ["HOMEPATH"]
            os.chdir(path)

            os.chdir("Desktop\\gpa")


            gr=['A','B','C','D','E','S']
            


            xl1=pd.read_excel('sem1.xlsx',sheet_name="Sheet1",skiprows=2)
            xl2=pd.read_excel('sem2.xlsx',sheet_name="Sheet1",skiprows=2)
            xl3=pd.read_excel('sem3.xlsx',sheet_name="Sheet1",skiprows=2)
            xl4=pd.read_excel('sem4.xlsx',sheet_name="Sheet1",skiprows=2)
            xl5=pd.read_excel('sem5.xlsx',sheet_name="Sheet1",skiprows=2)
            xl6=pd.read_excel('sem6.xlsx',sheet_name="Sheet1",skiprows=2)
            xl7=pd.read_excel('sem7.xlsx',sheet_name="Sheet1",skiprows=2)
            xl8=pd.read_excel('sem8.xlsx',sheet_name="Sheet1",skiprows=2)
            ar1=pd.DataFrame()
            if (tn2=="C"):
                
                if (tn==1):

                    xl1=xl1.drop(xl1.index[-8:])
                    


                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                   

                elif (tn==2):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    

                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                  


                elif (tn==3):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    xl3=xl3.drop(xl3.index[-8:])
                    

                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                    ar1[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    


                elif (tn==4):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    xl3=xl3.drop(xl3.index[-8:])
                    xl4=xl4.drop(xl4.index[-8:])
                    


                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                    ar1[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    ar1[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]=xl4[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]
                   



                elif (tn==5):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    xl3=xl3.drop(xl3.index[-8:])
                    xl4=xl4.drop(xl4.index[-8:])
                    xl5=xl5.drop(xl5.index[-8:])
                   

                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                    ar1[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    ar1[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]=xl4[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]
                    ar1[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]=xl5[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]
                    



                elif (tn==6):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    xl3=xl3.drop(xl3.index[-8:])
                    xl4=xl4.drop(xl4.index[-8:])
                    xl5=xl5.drop(xl5.index[-8:])
                    xl6=xl6.drop(xl6.index[-8:])
                    


                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                    ar1[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    ar1[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]=xl4[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]
                    ar1[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]=xl5[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]
                    ar1[['CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]=xl6[['CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]
                    



                elif (tn==7):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    xl3=xl3.drop(xl3.index[-8:])
                    xl4=xl4.drop(xl4.index[-8:])
                    xl5=xl5.drop(xl5.index[-8:])
                    xl6=xl6.drop(xl6.index[-8:])
                    xl7=xl7.drop(xl7.index[-8:])
                    


                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                    ar1[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    ar1[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]=xl4[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]
                    ar1[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]=xl5[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]
                    ar1[['CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]=xl6[['CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]
                    ar1[['CS6701', 'CS6702', 'CS6703', 'CS6704', 'ELEC-2','ELEC-3', 'CS6711', 'CS6712']]=xl7[['CS6701', 'CS6702', 'CS6703', 'CS6704', 'ELEC-2','ELEC-3', 'CS6711', 'CS6712']]
                   




                elif (tn==8):

                    xl1=xl1.drop(xl1.index[-8:])
                    xl2=xl2.drop(xl2.index[-8:])
                    xl3=xl3.drop(xl3.index[-8:])
                    xl4=xl4.drop(xl4.index[-8:])
                    xl5=xl5.drop(xl5.index[-8:])
                    xl6=xl6.drop(xl6.index[-8:])
                    xl7=xl7.drop(xl7.index[-8:])
                    xl8=xl8.drop(xl8.index[-8:])



                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]
                    ar1[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[[ 'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                    ar1[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    ar1[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]=xl4[['MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]
                    ar1[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]=xl5[['MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]
                    ar1[['CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]=xl6[['CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]
                    ar1[['CS6701', 'CS6702', 'CS6703', 'CS6704', 'ELEC-2','ELEC-3', 'CS6711', 'CS6712']]=xl7[['CS6701', 'CS6702', 'CS6703', 'CS6704', 'ELEC-2','ELEC-3', 'CS6711', 'CS6712']]
                    ar1[['CS6801', 'ELEC-4', 'ELEC-5', 'CS6811']]=xl8[['CS6801', 'ELEC-4', 'ELEC-5', 'CS6811']]


            elif(tn2=="S"):
                
                if (tn==1):

                    
                    xl1=xl1.drop(xl1.index[-8:])
                    
                    ar1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]=xl1[['REGISTER_NO','NAME',
                             'HS6151', 'MA6151', 'PH6151', 'CY6151', 'GE6151','GE6152', 'GE6161', 'GE6162', 'GE6163']]

                                   

                elif (tn==2):

                    
                    xl2=xl2.drop(xl2.index[-8:])
                    

                    
                    ar1[[ 'REGISTER_NO','NAME',
                          'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]=xl2[['REGISTER_NO','NAME',
                                                                                                                        'HS6251', 'MA6251', 'PH6251', 'CY6251','CS6201','CS6202', 'GE6262', 'CS6211', 'CS6212']]
                  


                elif (tn==3):

                   
                    xl3=xl3.drop(xl3.index[-8:])
                    
                    ar1[['REGISTER_NO','NAME',
                         'MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]=xl3[['REGISTER_NO','NAME',
                                                                                                              'MA6351', 'CS6301', 'CS6302', 'CS6303', 'CS6304','GE6351', 'CS6311', 'CS6312']]
                    


                elif (tn==4):

                    xl4=xl4.drop(xl4.index[-8:])
                    

                    ar1[['REGISTER_NO','NAME','MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]=xl4[['REGISTER_NO','NAME',
                                                                                                                                             'MA6453', 'CS6551', 'CS6401', 'CS6402', 'EC6504','CS6403', 'CS6411', 'CS6412', 'CS6413']]
                   



                elif (tn==5):
                    
                    xl5=xl5.drop(xl5.index[-8:])
                   
                    ar1[['REGISTER_NO','NAME','MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]=xl5[['REGISTER_NO','NAME',
                                                                                                                                   'MA6566', 'CS6501', 'CS6502', 'CS6503', 'CS6504','CS6511', 'CS6512', 'CS6513']]
                    



                elif (tn==6):

                    xl6=xl6.drop(xl6.index[-8:])
                    ar1[['REGISTER_NO','NAME','CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]=xl6[['REGISTER_NO','NAME',
                                                                                                                                           'CS6601', 'IT6601', 'CS6660', 'IT6502', 'CS6659','ELEC', 'CS6611', 'CS6612', 'GE6674']]
                    



                elif (tn==7):
                    
                    xl7=xl7.drop(xl7.index[-8:])
                    
                    ar1[['REGISTER_NO','NAME','CS6701', 'CS6702', 'CS6703', 'CS6704', 'ELEC-2','ELEC-3', 'CS6711', 'CS6712']]=xl7[['REGISTER_NO','NAME',
                                                                                                                                   'CS6701', 'CS6702', 'CS6703', 'CS6704', 'ELEC-2','ELEC-3', 'CS6711', 'CS6712']]
                   




                elif (tn==8):
                    
                    xl8=xl8.drop(xl8.index[-8:])


                    ar1[['REGISTER_NO','NAME','CS6801', 'ELEC-4', 'ELEC-5', 'CS6811']]=xl8[['REGISTER_NO','NAME',
                                                                                            'CS6801', 'ELEC-4', 'ELEC-5', 'CS6811']]

            
            #print(ar1)
            y=ar1.shape
            #print(y)
            colnam=ar1.columns

            for i in range(0,y[0]):
                for j in colnam[2:]:
                    if (ar1.at[i,j]) in gr:
                        ar1.at[i,j]=None
                        
                    elif(ar1.at[i,j]) =='WH':
                        ar1.at[i,j]='WH'

                    elif(ar1.at[i,j]) =='UA':
                        ar1.at[i,j]='√/*'

                    elif(ar1.at[i,j]) =='U':
                        ar1.at[i,j]='√'
                    


            #print(ar1)

            #-----------------------------subwise--------------------------------------------------------------------------
            
            subwise=pd.DataFrame()
            ar9=[]

            for j in colnam[2:]:
                w1=0
                for i in range(0,y[0]):
                    if (ar1.at[i,j]) !=None:
                        w1=w1+1
                ar9.append(w1)
            #print(ar9)
            dum=[None,None]
            ar9=dum+ar9

            #print(ar9)

            col=ar1.columns

            s=pd.Series(ar9,index=col)         
            
            #print(s)

            ar1=ar1.append(s,ignore_index=True)
            #print(ar1)            
            

            #---------------stuwise-------------------------------------------------
                
            stuwise=pd.DataFrame()
            ar=[]
            

            for i in range(0,y[0]):
                w=0
                
                for j in colnam[2:]:
                    if (ar1.at[i,j]) !=None:
                        w=w+1
                ar.append(w)
            #print(ar)

            stuwise['NO_OF_FAILS']=ar
                
            #print(stuwise)

            
            ar1['NO_OF_FAILS']=stuwise['NO_OF_FAILS']

            #print(ar1)

                                                                                    
            writ=pd.ExcelWriter('tick.xlsx',engine='xlsxwriter')
            
            ar1.to_excel(writ,sheet_name="Sheet1",startrow=3)

            msg.showinfo("SUCCESS","TICK FORMAT CREATED!!")

            writ.save()

        bes5=ttk.Button(end_frame,text="GENERATE",command=yg)
        bes5.grid(row=10,column=0,pady=10,padx=10)

        
#-------------------------------------------------------imports-------------------------    

    def s1(self):
        import semester_1

    def s2(self):
        import semester_2

    def s3(self):
        import semester_3

    def s4(self):
        import semester_4

    def s5(self):
        import semester_5

    def s6(self):
        import semester_6

    def s7(self):
        import semester_7
        
    def s8(self):
        import semester_8

path=os.environ["HOMEPATH"]
os.chdir(path)
os.chdir("Desktop\\gpa")

scr=tk.Tk()
scr.title('RESULT ANALYZER (REG_13)')
scr.config(bg="#00CED1")
scr.iconbitmap(r'icofile.ico')
home=main(scr)
scr.mainloop()
